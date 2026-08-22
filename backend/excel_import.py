from __future__ import annotations

import io
import json
import re
from typing import Any

from openpyxl import load_workbook

from . import llm, storage

OVERWRITE_FIELDS = (
    "name",
    "phone",
    "company",
    "title",
    "industry",
    "company_size",
    "notes",
    "source",
)


def normalize_phone(phone: str) -> str:
    digits = re.sub(r"\D", "", str(phone or ""))
    if digits.startswith("00") and len(digits) > 11:
        digits = digits[2:]
    if digits.startswith("86") and len(digits) >= 13:
        rest = digits[2:]
        if len(rest) == 11 and rest.startswith("1"):
            digits = rest
    return digits


def _incoming_from_row(item: dict[str, Any]) -> dict[str, str] | None:
    company = str(item.get("company") or "").strip()
    if not company:
        return None
    return {
        "name": str(item.get("name") or "").strip() or "客户",
        "phone": str(item.get("phone") or "").strip(),
        "company": company,
        "title": str(item.get("title") or "").strip(),
        "industry": str(item.get("industry") or "").strip(),
        "company_size": str(item.get("company_size") or "").strip(),
        "source": "Excel导入",
        "notes": str(item.get("notes") or "").strip(),
    }


def _lead_index_by_phone(leads: list[dict[str, Any]]) -> dict[str, dict[str, Any]]:
    index: dict[str, dict[str, Any]] = {}
    for lead in leads:
        key = normalize_phone(lead.get("phone") or "")
        if not key:
            continue
        prev = index.get(key)
        if prev is None or str(lead.get("updated_at") or "") >= str(prev.get("updated_at") or ""):
            index[key] = lead
    return index


def _public_existing(lead: dict[str, Any]) -> dict[str, Any]:
    return {
        "id": lead.get("id"),
        "name": lead.get("name") or "",
        "phone": lead.get("phone") or "",
        "company": lead.get("company") or "",
        "title": lead.get("title") or "",
        "industry": lead.get("industry") or "",
        "notes": lead.get("notes") or "",
        "status": lead.get("status") or "",
        "last_tier": lead.get("last_tier") or "",
    }


def apply_overwrite(existing_id: str, incoming: dict[str, Any]) -> dict[str, Any]:
    existing = storage.get_item("leads", existing_id)
    if not existing:
        raise ValueError("线索不存在")
    incoming_phone = normalize_phone(incoming.get("phone") or "")
    existing_phone = normalize_phone(existing.get("phone") or "")
    if incoming_phone and existing_phone and incoming_phone != existing_phone:
        raise ValueError("手机号与已有客户不一致，无法覆盖")
    updated = {**existing}
    for field in OVERWRITE_FIELDS:
        if field in incoming:
            value = incoming.get(field)
            updated[field] = "" if value is None else str(value).strip()
    if not updated.get("name"):
        updated["name"] = "客户"
    if not updated.get("company"):
        raise ValueError("公司名称不能为空")
    updated["id"] = existing_id
    updated["source"] = updated.get("source") or "Excel导入"
    return storage.upsert_item("leads", updated)


def workbook_to_preview(content: bytes, max_rows: int = 40) -> dict[str, Any]:
    wb = load_workbook(io.BytesIO(content), data_only=True)
    sheet = wb.active
    rows: list[list[Any]] = []
    for i, row in enumerate(sheet.iter_rows(values_only=True)):
        if i >= max_rows:
            break
        rows.append([("" if c is None else str(c).strip()) for c in row])
    rows = [r for r in rows if any(x for x in r)]
    headers = rows[0] if rows else []
    return {
        "sheet": sheet.title,
        "headers": headers,
        "row_count": len(rows) - 1 if rows else 0,
        "sample_rows": rows[:15],
        "all_rows": rows,
    }


def parse_leads_with_llm(content: bytes, filename: str = "upload.xlsx") -> dict[str, Any]:
    preview = workbook_to_preview(content)
    rows = preview["all_rows"]
    compact = rows[:80]
    system = (
        "你是肯耐珂萨（Kenexa）销售助理，负责把任意格式的客户 Excel 解析成标准线索。"
        "Excel 没有固定表头，请根据语义识别列：姓名/姓氏、电话/手机、公司/企业、职位、行业、备注等。"
        "输出 JSON：{leads:[{name,phone,company,title,industry,company_size,source,notes}], "
        "mapping_notes:string, skipped_rows:number}。"
        "name 可只保留姓或称呼；phone 尽量规范化为数字；company 必填，缺公司的行放入跳过。"
        "source 固定为「Excel导入」。不要编造电话或不存在的公司。"
    )
    user = (
        f"文件名：{filename}\n工作表：{preview['sheet']}\n"
        f"前若干行（含表头可能）：\n{json.dumps(compact, ensure_ascii=False)}\n"
        "请解析为线索 JSON。"
    )
    data = llm.chat_json(system=system, user=user, temperature=0.1, max_tokens=4000)
    leads_in = data.get("leads") or []

    incoming_rows: list[dict[str, str]] = []
    file_dupes = 0
    seen_in_file: dict[str, int] = {}
    for item in leads_in:
        incoming = _incoming_from_row(item)
        if not incoming:
            continue
        key = normalize_phone(incoming["phone"])
        if key:
            if key in seen_in_file:
                file_dupes += 1
                incoming_rows[seen_in_file[key]] = incoming
                continue
            seen_in_file[key] = len(incoming_rows)
        incoming_rows.append(incoming)

    existing_by_phone = _lead_index_by_phone(storage.list_items("leads"))
    saved: list[dict[str, Any]] = []
    conflicts: list[dict[str, Any]] = []
    for incoming in incoming_rows:
        key = normalize_phone(incoming["phone"])
        existing = existing_by_phone.get(key) if key else None
        if existing:
            conflicts.append(
                {
                    "phone": incoming["phone"] or existing.get("phone") or "",
                    "phone_key": key,
                    "existing": _public_existing(existing),
                    "incoming": incoming,
                }
            )
            continue
        lead = {
            "id": storage.new_id("L"),
            "status": "待分析",
            "workflow_step": 1,
            **incoming,
        }
        saved.append(storage.upsert_item("leads", lead))
        if key:
            existing_by_phone[key] = lead

    notes = str(data.get("mapping_notes") or "").strip()
    extras = []
    if file_dupes:
        extras.append(f"同一文件内重复手机号已合并 {file_dupes} 条（保留后出现的一行）")
    if conflicts:
        extras.append(f"发现 {len(conflicts)} 个已有手机号，请确认是否覆盖")
    mapping_notes = "；".join([p for p in (notes, *extras) if p])

    return {
        "imported": len(saved),
        "skipped": int(data.get("skipped_rows") or 0) + file_dupes,
        "conflicts": conflicts,
        "mapping_notes": mapping_notes,
        "preview": {
            "sheet": preview["sheet"],
            "headers": preview["headers"],
            "row_count": preview["row_count"],
            "sample_rows": preview["sample_rows"],
        },
        "leads": saved,
    }
